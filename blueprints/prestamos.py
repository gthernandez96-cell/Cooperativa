from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file, current_app, Response
import math
import json
import os
import csv
from io import BytesIO, StringIO
from datetime import date, datetime, timedelta
from werkzeug.utils import secure_filename

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from utils.db import get_db, db_fetchone, db_fetchall, db_execute, db_insert_and_get_id, db_executemany, ensure_required_configurations, ensure_default_prestamo_categories, ensure_system_settings, ensure_module_settings, set_system_setting, get_system_setting, _is_postgres_connection
from utils.decorators import login_required, permission_required
from utils.nombres import preparar_datos_socio, construir_nombre_completo, construir_apellido_completo, validar_dpi, resumen_beneficiarios
from utils.financial import calcular_resumen_prestamo, generar_calendario_prestamo, normalizar_fecha_referencia, calcular_total_cuotas_prestamo, calcular_proximo_pago, obtener_dias_frecuencia, fecha_quincenal_mas_cercana, calcular_alerta_prestamo
from utils.prestamos import obtener_cartera_con_alertas, cargar_contexto_nuevo_prestamo
from utils.helpers import log_auditoria_evento, periodo_cerrado, generar_numero_comprobante, validar_pago_frecuencia, obtener_mensaje_validacion_frecuencia, calcular_bono_14, calcular_aguinaldo, formatear_fecha_dmy
from config import SYSTEM_SETTINGS_DEFAULTS, PRESTAMO_SETTINGS_DEFAULTS, DEFAULT_COOPERATIVA_NOMBRE

bp = Blueprint('prestamos', __name__)

def renderizar_finiquito_prestamo(prestamo, plantilla):
    calendario = prestamo.get('calendario') or []
    fecha_primer_pago = calendario[0]['fecha_programada'] if calendario else '—'
    fecha_ultima_cuota = calendario[-1]['fecha_programada'] if calendario else '—'
    contexto = {
        'cooperativa_nombre': prestamo.get('cooperativa_nombre') or DEFAULT_COOPERATIVA_NOMBRE,
        'prestamo_numero': prestamo.get('numero') or '',
        'socio_nombre': prestamo.get('nombre_socio') or '',
        'socio_codigo': prestamo.get('socio_codigo') or '',
        'categoria_nombre': prestamo.get('categoria_nombre') or 'General',
        'monto_aprobado': f"{float(prestamo.get('monto_aprobado') or prestamo.get('monto_solicitado') or 0):,.2f}",
        'cuota': f"{float(prestamo.get('cuota_mensual') or 0):,.2f}",
        'frecuencia': prestamo.get('frecuencia') or 'Quincenal',
        'fecha_aprobacion': prestamo.get('fecha_aprobacion') or date.today().isoformat(),
        'fecha_primer_pago': fecha_primer_pago,
        'fecha_ultima_cuota': fecha_ultima_cuota,
        'total_cuotas': prestamo.get('total_cuotas') or len(calendario),
        'estado': prestamo.get('estado') or '',
        'desembolso_tipo': prestamo.get('desembolso_tipo') or 'No definido',
        'desembolso_referencia': prestamo.get('desembolso_referencia') or 'Sin referencia',
    }

    try:
        return (plantilla or SYSTEM_SETTINGS_DEFAULTS['prestamo_finiquito_texto']).format(**contexto)
    except KeyError:
        return SYSTEM_SETTINGS_DEFAULTS['prestamo_finiquito_texto'].format(**contexto)



@bp.route('/prestamos')
def prestamos():
    rows = obtener_cartera_con_alertas()
    q = request.args.get('q', '').strip()
    estado_filtro = request.args.get('estado', '').strip().lower()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    ordenar_por = request.args.get('ordenar_por', 'fecha_solicitud').strip().lower()
    direccion = request.args.get('direccion', 'desc').strip().lower()
    vista = request.args.get('vista', 'activos').strip().lower()
    vistas_validas = {'activos', 'pagados', 'pendientes'}
    if vista not in vistas_validas:
        vista = 'activos'
    if direccion not in ('asc', 'desc'):
        direccion = 'desc'

    if q:
        q_lower = q.lower()
        filtrados = []
        for item in rows:
            texto_busqueda = ' '.join([
                str(item.get('numero') or ''),
                str(item.get('nombre_socio') or ''),
                str(item.get('socio_codigo') or ''),
                str(item.get('categoria_nombre') or ''),
                str(item.get('estado') or ''),
                str(item.get('desembolso_referencia') or ''),
            ]).lower()
            if q_lower in texto_busqueda:
                filtrados.append(item)
        rows = filtrados

    if estado_filtro:
        rows = [item for item in rows if (item.get('estado') or '').lower() == estado_filtro]

    if fecha_desde:
        rows = [
            item for item in rows
            if (item.get('fecha_solicitud') or '')[:10] >= fecha_desde
        ]

    if fecha_hasta:
        rows = [
            item for item in rows
            if (item.get('fecha_solicitud') or '')[:10] <= fecha_hasta
        ]

    conteos = {
        'activos': sum(1 for item in rows if item.get('estado') == 'aprobado' and float(item.get('saldo_pendiente') or 0) > 0),
        'pagados': sum(1 for item in rows if item.get('estado') == 'pagado' or (item.get('estado') == 'aprobado' and float(item.get('saldo_pendiente') or 0) <= 0)),
        'pendientes': sum(1 for item in rows if item.get('estado') == 'pendiente'),
    }

    if vista == 'activos':
        rows = [item for item in rows if item.get('estado') == 'aprobado' and float(item.get('saldo_pendiente') or 0) > 0]
        subtitulo = 'Préstamos aprobados con saldo pendiente'
    elif vista == 'pagados':
        rows = [item for item in rows if item.get('estado') == 'pagado' or (item.get('estado') == 'aprobado' and float(item.get('saldo_pendiente') or 0) <= 0)]
        subtitulo = 'Préstamos cancelados o liquidados'
    else:
        rows = [item for item in rows if item.get('estado') == 'pendiente']
        subtitulo = 'Solicitudes pendientes de aprobación'

    reverse = direccion == 'desc'
    if ordenar_por == 'saldo_pendiente':
        rows = sorted(rows, key=lambda item: float(item.get('saldo_pendiente') or 0), reverse=reverse)
    elif ordenar_por == 'cuota_mensual':
        rows = sorted(rows, key=lambda item: float(item.get('cuota_mensual') or 0), reverse=reverse)
    elif ordenar_por == 'monto_solicitado':
        rows = sorted(rows, key=lambda item: float(item.get('monto_solicitado') or 0), reverse=reverse)
    elif ordenar_por == 'nombre_socio':
        rows = sorted(rows, key=lambda item: str(item.get('nombre_socio') or '').lower(), reverse=reverse)
    else:
        ordenar_por = 'fecha_solicitud'
        rows = sorted(rows, key=lambda item: str(item.get('fecha_solicitud') or ''), reverse=reverse)

    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = min(100, max(10, int(request.args.get('per_page', 25) or 25)))
    total = len(rows)
    total_pages = max(1, math.ceil(total / per_page))

    # ── Totales de cartera (sobre todos los filtrados, sin paginar) ───────────
    totales_cartera = {
        'monto': sum(float(r.get('monto_solicitado') or 0) for r in rows),
        'saldo': sum(float(r.get('saldo_pendiente') or 0) for r in rows),
        'cuota': sum(float(r.get('cuota_mensual') or 0) for r in rows),
    }

    # ── Exportación CSV ───────────────────────────────────────────────────────
    export = request.args.get('export', '').strip().lower()
    if export == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Número', 'Socio', 'Código', 'Categoría', 'Desembolso', 'Monto', 'Tasa %', 'Plazo', 'Cuota', 'Saldo', 'Estado', 'Semáforo'])
        for r in rows:
            writer.writerow([
                r.get('numero', ''), r.get('nombre_socio', ''), r.get('socio_codigo', ''),
                r.get('categoria_nombre', ''), r.get('desembolso_tipo', ''),
                r.get('monto_solicitado', 0), r.get('tasa_interes', 0), r.get('plazo_meses', 0),
                r.get('cuota_mensual', 0), r.get('saldo_pendiente', 0),
                r.get('estado', ''), r.get('semaforo', ''),
            ])
        filename = f"cartera_prestamos_{vista}_{date.today().isoformat()}.csv"
        return Response(output.getvalue(), mimetype='text/csv; charset=utf-8',
                        headers={'Content-Disposition': f'attachment; filename={filename}'})

    if export == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cartera'
        headers = ['Número', 'Socio', 'Código', 'Categoría', 'Monto', 'Tasa %', 'Plazo', 'Cuota', 'Saldo', 'Estado', 'Semáforo']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='003B74')
            cell.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append([
                r.get('numero', ''), r.get('nombre_socio', ''), r.get('socio_codigo', ''),
                r.get('categoria_nombre', ''), float(r.get('monto_solicitado') or 0),
                float(r.get('tasa_interes') or 0), r.get('plazo_meses', 0),
                float(r.get('cuota_mensual') or 0), float(r.get('saldo_pendiente') or 0),
                r.get('estado', ''), r.get('semaforo', ''),
            ])
        mem = BytesIO()
        wb.save(mem)
        mem.seek(0)
        filename = f"cartera_prestamos_{vista}_{date.today().isoformat()}.xlsx"
        return send_file(mem, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Paginación (solo para HTML) ───────────────────────────────────────────
    start_idx = (page - 1) * per_page
    rows = rows[start_idx:start_idx + per_page]

    return render_template(
        'prestamos.html',
        prestamos=rows,
        vista=vista,
        conteos=conteos,
        subtitulo=subtitulo,
        totales_cartera=totales_cartera,
        q=q,
        estado_filtro=estado_filtro,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        ordenar_por=ordenar_por,
        direccion=direccion,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@bp.route('/prestamos/<int:pid>')
@login_required()
def detalle_prestamo(pid):
    conn = get_db()
    prestamo = db_fetchone(
        conn,
        '''
        SELECT p.*, s.codigo AS socio_codigo, s.nombre || ' ' || s.apellido AS nombre_socio,
               s.frecuencia, pc.nombre AS categoria_nombre,
               COALESCE(pp.cuotas_pagadas, 0) AS cuotas_pagadas,
               COALESCE(pp.monto_pagado, 0) AS monto_pagado,
               pp.ultimo_pago
        FROM prestamos p
        JOIN socios s ON s.id = p.socio_id
        LEFT JOIN prestamo_categorias pc ON pc.id = p.categoria_id
        LEFT JOIN (
            SELECT prestamo_id,
                   COUNT(*) AS cuotas_pagadas,
                   SUM(monto) AS monto_pagado,
                   MAX(fecha) AS ultimo_pago
            FROM pagos_prestamo
            GROUP BY prestamo_id
        ) pp ON p.id = pp.prestamo_id
        WHERE p.id=?
        ''',
        [pid]
    )

    if not prestamo:
        conn.close()
        flash('Prestamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    prestamo = dict(prestamo)
    if (prestamo.get('estado') or '').lower() == 'amortizado':
        prestamo['estado'] = 'pagado'

    pagos = db_fetchall(
        conn,
        '''
        SELECT *
        FROM pagos_prestamo
        WHERE prestamo_id=?
        ORDER BY date(fecha) DESC, id DESC
        ''',
        [pid]
    )
    pagos = [dict(row) for row in pagos]

    # Si este préstamo fue amortizado por uno o más préstamos nuevos,
    # se muestran como movimientos dentro del historial de cuotas pagadas.
    prestamos_pagadores = db_fetchall(
        conn,
        '''
        SELECT numero,
               fecha_aprobacion,
               fecha_solicitud,
               COALESCE(monto_amortizado, 0) AS monto_amortizado,
               COALESCE(monto_aprobado, monto_solicitado, 0) AS total_prestamo
        FROM prestamos
        WHERE refinanciado_de=?
          AND COALESCE(monto_amortizado, 0) > 0
        ORDER BY date(COALESCE(fecha_aprobacion, fecha_solicitud)) DESC, id DESC
        ''',
        [pid]
    )
    for prestamo_pagador in prestamos_pagadores:
        pagos.append({
            'id': None,
            'fecha': prestamo_pagador['fecha_aprobacion'] or prestamo_pagador['fecha_solicitud'],
            'monto': float(prestamo_pagador['monto_amortizado'] or 0),
            'capital': float(prestamo_pagador['monto_amortizado'] or 0),
            'interes': 0.0,
            'saldo_restante': float(prestamo.get('saldo_pendiente') or 0),
            'boleta_deposito': 'Amortización',
            'numero_comprobante': None,
            'es_amortizacion': True,
            'prestamo_origen_numero': prestamo_pagador['numero'] or '—',
            'total_prestamo_origen': float(prestamo_pagador['total_prestamo'] or 0),
        })

    if prestamo.get('refinanciado_de') and float(prestamo.get('monto_amortizado') or 0) > 0:
        prestamo_amortizado = db_fetchone(
            conn,
            '''
            SELECT numero,
                   COALESCE(monto_aprobado, monto_solicitado, 0) AS total_prestamo
            FROM prestamos
            WHERE id=?
            ''',
            [prestamo['refinanciado_de']]
        )

        pagos.insert(0, {
            'id': None,
            'fecha': prestamo.get('fecha_aprobacion') or prestamo.get('fecha_solicitud'),
            'monto': float(prestamo.get('monto_amortizado') or 0),
            'capital': float(prestamo.get('monto_amortizado') or 0),
            'interes': 0.0,
            'saldo_restante': float(prestamo.get('saldo_pendiente') or 0),
            'boleta_deposito': 'Amortización',
            'numero_comprobante': None,
            'es_amortizacion': True,
            'prestamo_amortizado_numero': prestamo_amortizado['numero'] if prestamo_amortizado else '—',
            'total_prestamo_amortizado': float(prestamo_amortizado['total_prestamo'] or 0) if prestamo_amortizado else 0.0,
        })

    pagos = sorted(
        pagos,
        key=lambda item: (str(item.get('fecha') or ''), int(item.get('id') or 0)),
        reverse=True,
    )

    calendario = db_fetchall(
        conn,
        '''
        SELECT numero_cuota, fecha_programada, monto_programado, estado
        FROM prestamo_calendario_pagos
        WHERE prestamo_id=?
        ORDER BY numero_cuota
        ''',
        [pid]
    )

    conn.close()

    prestamo_dict = dict(prestamo)
    prestamo_dict['total_cuotas'] = calcular_total_cuotas_prestamo(
        prestamo_dict.get('plazo_meses'),
        prestamo_dict.get('frecuencia')
    )
    if calendario:
        prestamo_dict['cuotas_pagadas_calendario'] = sum(1 for cuota in calendario if (cuota['estado'] or '').lower() == 'pagado')
    else:
        prestamo_dict['cuotas_pagadas_calendario'] = prestamo_dict.get('cuotas_pagadas') or 0

    return render_template(
        'detalle_prestamo.html',
        prestamo=prestamo_dict,
        pagos=pagos,
        calendario=calendario,
    )


@bp.route('/prestamos/nuevo', methods=['GET', 'POST'])
def nuevo_prestamo():
    conn = get_db()

    if request.method == 'POST':
        socio_id = request.form.get('socio_id', '').strip()
        categoria_id = request.form.get('categoria_id', '').strip()
        # Lista de IDs de préstamos seleccionados para amortizar (checkboxes)
        prestamos_a_amortizar_ids = request.form.getlist('prestamos_a_amortizar[]')
        forma_desembolso = request.form.get('forma_desembolso', 'cheque').strip() or 'cheque'

        if not socio_id:
            conn.close()
            flash('Debe seleccionar un asociado válido.', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo'))

        socio = db_fetchone(
            conn,
            "SELECT id, frecuencia, banco_nombre, banco_tipo_cuenta, banco_numero_cuenta FROM socios WHERE id=?",
            [socio_id]
        )
        if not socio:
            conn.close()
            flash('Debe seleccionar un asociado válido.', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo'))

        categoria = None
        if categoria_id:
            categoria = db_fetchone(
                conn,
                "SELECT id FROM prestamo_categorias WHERE id=? AND estado='activo'",
                [categoria_id]
            )
        if not categoria:
            conn.close()
            flash('Debe seleccionar una categoria de prestamo válida.', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        try:
            monto = float(request.form.get('monto', 0) or 0)
            tasa = float(request.form.get('tasa', 0) or 0)
            plazo = int(request.form.get('plazo', 0) or 0)
        except (TypeError, ValueError):
            conn.close()
            flash('Los datos del préstamo no son válidos.', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        if monto <= 0 or tasa <= 0 or plazo <= 0:
            conn.close()
            flash('Debe ingresar monto, tasa y plazo válidos.', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        banco_tipo = ''
        banco_numero = ''
        if forma_desembolso == 'deposito':
            banco_tipo = (socio['banco_tipo_cuenta'] or '').strip()
            banco_numero = (socio['banco_numero_cuenta'] or '').strip()
            banco_nombre = (socio['banco_nombre'] or '').strip()
            if not banco_nombre or not banco_tipo or not banco_numero:
                conn.close()
                flash('El asociado no tiene la información bancaria completa para desembolso por deposito.', 'danger')
                return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        # Validar si tiene algún préstamo pendiente de aprobación/desembolso
        prestamo_pendiente = db_fetchone(
            conn,
            "SELECT numero FROM prestamos WHERE socio_id=? AND estado='pendiente'",
            [socio_id]
        )
        if prestamo_pendiente:
            conn.close()
            flash(f'El asociado ya tiene una solicitud pendiente ({prestamo_pendiente["numero"]}). Debe ser resuelta antes de crear una nueva.', 'warning')
            return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        # Obtener todos los préstamos vigentes del asociado
        prestamos_vigentes = db_fetchall(
            conn,
            '''
            SELECT id, numero, categoria_id,
                   COALESCE(saldo_pendiente, monto_aprobado, monto_solicitado, 0) AS saldo_vigente
            FROM prestamos
            WHERE socio_id=? AND estado IN ('pendiente', 'aprobado')
            ORDER BY date(fecha_solicitud) DESC, id DESC
            ''',
            [socio_id]
        )

        # Verificar si existe algún préstamo de la MISMA categoría
        ids_misma_categoria = [
            str(p['id']) for p in prestamos_vigentes
            if str(p['categoria_id']) == str(categoria_id)
        ]

        if ids_misma_categoria:
            # Hay préstamo(s) de misma categoría → amortización OBLIGATORIA
            seleccionados_misma_cat = [
                pid for pid in prestamos_a_amortizar_ids
                if pid in ids_misma_categoria
            ]
            if not seleccionados_misma_cat:
                conn.close()
                flash(
                    'El asociado ya tiene un préstamo vigente de la misma categoría. '
                    'Debe seleccionarlo para amortizarlo antes de otorgar uno nuevo.',
                    'danger'
                )
                return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

        resumen = calcular_resumen_prestamo(monto, tasa, plazo, socio['frecuencia'])

        try:
            count = db_fetchone(conn, "SELECT COUNT(*) FROM prestamos")[0]
            numero = f'PRE-{count+1:04d}'

            if prestamos_a_amortizar_ids:
                total_capital_amort = 0.0
                total_interes_amort = 0.0
                
                # Préstamo principal (refinanciado_de)
                p_principal_id = prestamos_a_amortizar_ids[0]
                p_principal = db_fetchone(
                    conn,
                    "SELECT id, numero, COALESCE(saldo_pendiente, monto_aprobado, monto_solicitado, 0) AS saldo_vigente FROM prestamos WHERE id=?",
                    [p_principal_id]
                )
                
                # Calcular total amortizado sumando todos los seleccionados
                for pid in prestamos_a_amortizar_ids:
                    try:
                        c_a = float(request.form.get(f'capital_amortizado_{pid}', 0) or 0)
                        i_a = float(request.form.get(f'interes_amortizado_{pid}', 0) or 0)
                    except (TypeError, ValueError):
                        c_a = 0.0
                        i_a = 0.0
                    
                    if c_a + i_a == 0:
                        # Si no se ingresaron valores, usar el saldo vigente del préstamo
                        p_v = db_fetchone(conn, "SELECT COALESCE(saldo_pendiente, monto_aprobado, monto_solicitado, 0) AS saldo_vigente FROM prestamos WHERE id=?", [pid])
                        c_a = float(p_v['saldo_vigente'] or 0)
                        i_a = 0.0
                    
                    total_capital_amort += c_a
                    total_interes_amort += i_a

                total_amortizado = total_capital_amort + total_interes_amort
                
                if total_amortizado > monto:
                    conn.close()
                    flash(f'El saldo total a amortizar (Q{total_amortizado:,.2f}) no puede ser mayor al monto solicitado (Q{monto:,.2f}).', 'danger')
                    return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))

                monto_desembolso = max(0, monto - total_amortizado)

                db_execute(
                    conn,
                    '''
                    INSERT INTO prestamos (
                        numero, socio_id, categoria_id, monto_solicitado, tasa_interes, plazo_meses,
                        cuota_mensual, fecha_solicitud, desembolso_tipo, banco_tipo_cuenta,
                        banco_numero_cuenta, refinanciado_de, monto_amortizado, monto_desembolso,
                        capital_amortizado, interes_amortizado
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ''',
                    (
                        numero, socio_id, categoria['id'], monto, tasa, plazo,
                        resumen['cuota'], date.today().isoformat(),
                        forma_desembolso, banco_tipo, banco_numero,
                        p_principal['id'], total_amortizado, monto_desembolso,
                        total_capital_amort, total_interes_amort,
                    )
                )
                numeros_amortizados = []
                # Si hay préstamos seleccionados, registrarlos y agregarlos a la lista
                for extra_id in prestamos_a_amortizar_ids:
                    try:
                        capital_e = float(request.form.get(f'capital_amortizado_{extra_id}', 0) or 0)
                        interes_e = float(request.form.get(f'interes_amortizado_{extra_id}', 0) or 0)
                    except (TypeError, ValueError):
                        capital_e = 0.0
                        interes_e = 0.0
                    
                    prestamo_extra = db_fetchone(
                        conn,
                        "SELECT numero FROM prestamos WHERE id=? AND socio_id=?",
                        [extra_id, socio_id]
                    )
                    if prestamo_extra:
                        numeros_amortizados.append(prestamo_extra['numero'])
                        # Solo auditar si NO es el principal (el principal ya se guarda en la tabla prestamos)
                        if str(extra_id) != str(p_principal['id']):
                            db_execute(
                                conn,
                                '''
                                INSERT INTO auditoria_eventos
                                    (modulo, entidad, entidad_id, accion, descripcion, usuario, fecha)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                                ''',
                                (
                                    'prestamos', 'prestamo', extra_id, 'amortizacion_adicional',
                                    f'Capital: Q{capital_e:,.2f} | Interés: Q{interes_e:,.2f} '
                                    f'| Préstamo nuevo: {numero}',
                                    session.get('username', 'sistema'),
                                    date.today().isoformat(),
                                )
                            )

                mensaje = (
                    f'Solicitud enviada. Si se aprueba, amortizará: '
                    f'{", ".join(numeros_amortizados)}. '
                    f'Desembolso estimado: Q{monto_desembolso:,.2f}'
                )
            else:
                db_execute(
                    conn,
                    '''
                    INSERT INTO prestamos (
                        numero, socio_id, categoria_id, monto_solicitado, tasa_interes, plazo_meses,
                        cuota_mensual, fecha_solicitud, desembolso_tipo, banco_tipo_cuenta,
                        banco_numero_cuenta, monto_desembolso
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    ''',
                    (
                        numero, socio_id, categoria['id'], monto, tasa, plazo,
                        resumen['cuota'], date.today().isoformat(),
                        forma_desembolso, banco_tipo, banco_numero, monto,
                    )
                )
                mensaje = 'Solicitud de préstamo enviada.'

            conn.commit()
            flash(mensaje, 'success')
            return redirect(url_for('prestamos.prestamos'))
        except Exception as e:
            conn.rollback()
            flash(f'Error: {e}', 'danger')
            return redirect(url_for('prestamos.nuevo_prestamo', socio_id=socio_id))
        finally:
            conn.close()

    socio_id_seleccionado = request.args.get('socio_id', '').strip()
    contexto = cargar_contexto_nuevo_prestamo(socio_id_seleccionado=socio_id_seleccionado)
    conn.close()
    return render_template('nuevo_prestamo.html', **contexto)

@bp.route('/api/calcular_prestamo_detalles')
@login_required()
def api_calcular_prestamo_detalles():
    socio_id = request.args.get('socio_id')
    if not socio_id:
        return jsonify({'error': 'socio_id requerido'}), 400
    
    conn = get_db()
    try:
        bono_14 = calcular_bono_14(socio_id, conn)
        aguinaldo = calcular_aguinaldo(socio_id, conn)
        
        # Obtener saldo total de ahorros
        ahorro_row = db_fetchone(conn, "SELECT SUM(saldo) FROM cuentas WHERE socio_id=? AND estado='activa'", [socio_id])
        saldo_ahorro = ahorro_row[0] if ahorro_row and ahorro_row[0] else 0.0

        return jsonify({
            'bono_14': float(bono_14) if bono_14 else 0.0,
            'aguinaldo': float(aguinaldo) if aguinaldo else 0.0,
            'saldo_ahorro': float(saldo_ahorro)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


def obtener_detalle_prestamo_aprobacion(conn, pid):
    prestamo = db_fetchone(
        conn,
        '''
        SELECT p.*, s.nombre || ' ' || s.apellido AS nombre_socio, s.codigo AS socio_codigo,
               s.frecuencia, pc.nombre AS categoria_nombre,
               s.banco_nombre, s.banco_tipo_cuenta, s.banco_numero_cuenta
        FROM prestamos p
        JOIN socios s ON s.id = p.socio_id
        LEFT JOIN prestamo_categorias pc ON pc.id = p.categoria_id
        WHERE p.id=?
        ''',
        [pid]
    )
    if not prestamo:
        return None

    calendario = db_fetchall(
        conn,
        '''
        SELECT numero_cuota, fecha_programada, monto_programado, estado
        FROM prestamo_calendario_pagos
        WHERE prestamo_id=?
        ORDER BY numero_cuota
        ''',
        [pid]
    )

    item = dict(prestamo)
    item['calendario'] = [dict(row) for row in calendario]
    item['total_cuotas'] = calcular_total_cuotas_prestamo(item.get('plazo_meses'), item.get('frecuencia'))
    return item

@bp.route('/prestamos/<int:pid>/aprobar', methods=['GET', 'POST'])
@login_required(role=('Administrador', 'Operador'))
@permission_required('prestamos.aprobar')
def aprobar_prestamo(pid):
    conn = get_db()
    prestamo = obtener_detalle_prestamo_aprobacion(conn, pid)
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    fecha_aprobacion = request.form.get('fecha_aprobacion') if request.method == 'POST' else (prestamo.get('fecha_aprobacion') or date.today().isoformat())
    fecha_primer_pago_default = prestamo.get('calendario', [{}])[0].get('fecha_programada') if prestamo.get('calendario') else None
    if not fecha_primer_pago_default:
        frecuencia_socio = prestamo.get('frecuencia') or 'Quincenal'
        fecha_ref = normalizar_fecha_referencia(fecha_aprobacion)
        if frecuencia_socio.lower() == 'catorcenal':
            fecha_primer_pago_default = (fecha_ref + timedelta(days=14)).isoformat()
        else:
            # Quincenal: próxima fecha válida (día 15 o último del mes)
            fecha_primer_pago_default = fecha_quincenal_mas_cercana(fecha_ref).isoformat()
    fecha_primer_pago = request.form.get('fecha_primer_pago') if request.method == 'POST' else fecha_primer_pago_default
    monto_aprobado = float(request.form.get('monto_aprobado', prestamo.get('monto_aprobado') or prestamo.get('monto_solicitado') or 0))
    desembolso_tipo = request.form.get('desembolso_tipo') if request.method == 'POST' else (prestamo.get('desembolso_tipo') or 'deposito')
    resumen = calcular_resumen_prestamo(monto_aprobado, prestamo['tasa_interes'], prestamo['plazo_meses'], prestamo['frecuencia'])
    if request.method == 'GET' and prestamo.get('calendario'):
        calendario_preview = generar_calendario_prestamo(
            prestamo['calendario'][0]['fecha_programada'],
            resumen['total_cuotas'], resumen['cuota'], prestamo['frecuencia'],
            monto=monto_aprobado, tasa_anual=prestamo['tasa_interes']
        )
    else:
        calendario_preview = generar_calendario_prestamo(
            fecha_primer_pago, resumen['total_cuotas'], resumen['cuota'], prestamo['frecuencia'],
            monto=monto_aprobado, tasa_anual=prestamo['tasa_interes']
        )

    def _render(**extra):
        return render_template(
            'aprobar_prestamo.html',
            prestamo=prestamo, resumen=resumen, calendario_preview=calendario_preview,
            fecha_aprobacion=fecha_aprobacion, fecha_primer_pago=fecha_primer_pago,
            desembolso_tipo=desembolso_tipo, **extra
        )

    if request.method == 'POST':
        if prestamo.get('estado') != 'pendiente':
            conn.close()
            flash('Solo se pueden aprobar solicitudes en estado pendiente.', 'warning')
            return redirect(url_for('prestamos.detalle_prestamo', pid=pid))

        if normalizar_fecha_referencia(fecha_primer_pago) <= normalizar_fecha_referencia(fecha_aprobacion):
            conn.close()
            flash('La primera fecha de pago debe ser posterior a la fecha de aprobación.', 'warning')
            return _render()

        if desembolso_tipo not in ('deposito', 'cheque'):
            conn.close()
            flash('Debe seleccionar una forma de desembolso válida.', 'warning')
            return _render()

        db_execute(
            conn,
            "UPDATE prestamos SET estado='aprobado', monto_aprobado=?, cuota_mensual=?, saldo_pendiente=?, fecha_aprobacion=?, desembolso_tipo=? WHERE id=?",
            [monto_aprobado, resumen['cuota'], monto_aprobado, fecha_aprobacion, desembolso_tipo, pid]
        )
        mensaje_aprobacion = 'Préstamo aprobado y calendario generado correctamente.'
        if prestamo.get('refinanciado_de'):
            prestamo_anterior = db_fetchone(
                conn,
                '''
                SELECT id, numero, COALESCE(saldo_pendiente, monto_aprobado, monto_solicitado, 0) AS saldo_vigente
                FROM prestamos
                WHERE id=?
                ''',
                [prestamo['refinanciado_de']]
            )
            if prestamo_anterior and (prestamo_anterior['saldo_vigente'] or 0) > 0:
                monto_amortizado = float(prestamo_anterior['saldo_vigente'] or 0)
                monto_desembolso = max(0, float(monto_aprobado or 0) - monto_amortizado)
                db_execute(
                    conn,
                    "UPDATE prestamos SET saldo_pendiente=0, estado='pagado' WHERE id=?",
                    [prestamo_anterior['id']]
                )
                db_execute(
                    conn,
                    "UPDATE prestamo_calendario_pagos SET estado='pagado' WHERE prestamo_id=? AND estado='pendiente'",
                    [prestamo_anterior['id']]
                )
                
                # BUSCAR AMORTIZACIONES ADICIONALES (Múltiples préstamos)
                amortizaciones_extras = db_fetchall(
                    conn,
                    "SELECT entidad_id FROM auditoria_eventos WHERE modulo='prestamos' AND accion='amortizacion_adicional' AND descripcion LIKE ?",
                    [f"%Préstamo nuevo: {prestamo['numero']}%"]
                )
                for extra in amortizaciones_extras:
                    db_execute(conn, "UPDATE prestamos SET saldo_pendiente=0, estado='pagado' WHERE id=?", [extra['entidad_id']])
                    db_execute(conn, "UPDATE prestamo_calendario_pagos SET estado='pagado' WHERE prestamo_id=? AND estado='pendiente'", [extra['entidad_id']])
                
                db_execute(
                    conn,
                    "UPDATE prestamos SET monto_amortizado=?, monto_desembolso=? WHERE id=?",
                    [monto_amortizado, monto_desembolso, pid]
                )
                db_execute(
                    conn,
                    '''
                    INSERT INTO auditoria_eventos (modulo, entidad, entidad_id, accion, descripcion, usuario, fecha)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''',
                    (
                        'prestamos', 'prestamo', prestamo_anterior['id'], 'amortizar',
                        f"Amortización de {prestamo_anterior['numero']} aplicada al aprobar el préstamo {prestamo['numero']}. Desembolso real Q{monto_desembolso:,.2f}",
                        session.get('username', 'sistema'),
                        date.today().isoformat(),
                    )
                )
                mensaje_aprobacion = f'Préstamo aprobado, calendario generado y amortizaciones aplicadas.'
        db_execute(conn, 'DELETE FROM prestamo_calendario_pagos WHERE prestamo_id=?', [pid])
        db_executemany(
            conn,
            '''
            INSERT INTO prestamo_calendario_pagos (prestamo_id, numero_cuota, fecha_programada, monto_programado, estado)
            VALUES (?, ?, ?, ?, 'pendiente')
            ''',
            [(pid, item['numero_cuota'], item['fecha_programada'], item['monto_programado']) for item in calendario_preview]
        )
        conn.commit()
        conn.close()
        flash(mensaje_aprobacion, 'success')
        return redirect(url_for('prestamos.aprobacion_exitosa', pid=pid))


    conn.close()
    return _render()

@bp.route('/prestamos/<int:pid>/aprobacion-exitosa')
@login_required()
def aprobacion_exitosa(pid):
    conn = get_db()
    prestamo = db_fetchone(
        conn,
        '''
        SELECT p.*, s.nombre || ' ' || s.apellido AS nombre_socio, s.codigo AS socio_codigo
        FROM prestamos p
        JOIN socios s ON s.id = p.socio_id
        WHERE p.id=?
        ''',
        [pid]
    )
    conn.close()
    if not prestamo:
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))
    
    return render_template('aprobacion_exitosa.html', prestamo=prestamo)


@bp.route('/prestamos/<int:pid>/no-procede', methods=['POST'])
@login_required()
def marcar_prestamo_no_procede(pid):
    conn = get_db()
    prestamo = db_fetchone(conn, "SELECT id, numero, estado FROM prestamos WHERE id=?", [pid])
    if not prestamo:
        conn.close()
        flash('Prestamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos', vista='pendientes'))

    if (prestamo['estado'] or '').lower() != 'pendiente':
        conn.close()
        flash('Solo se pueden marcar como no procede las solicitudes pendientes.', 'warning')
        return redirect(url_for('prestamos.prestamos', vista='pendientes'))

    db_execute(
        conn,
        "UPDATE prestamos SET estado='no_procede', saldo_pendiente=0 WHERE id=?",
        [pid]
    )
    conn.commit()
    conn.close()

    log_auditoria_evento(
        modulo='prestamos',
        entidad='prestamo',
        entidad_id=pid,
        accion='actualizar',
        descripcion=f'Solicitud de prestamo {prestamo["numero"]} marcada como no procede',
        datos={'estado': 'no_procede'}
    )

    flash('La solicitud se marco como no procede.', 'success')
    return redirect(url_for('prestamos.prestamos', vista='pendientes'))

@bp.route('/prestamos/editar/<int:pid>', methods=['GET', 'POST'])
@login_required(role='Administrador')
@permission_required('prestamos')
def editar_prestamo(pid):
    conn = get_db()
    prestamo = db_fetchone(conn, "SELECT * FROM prestamos WHERE id=?", [pid])
    
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))
    
    if (prestamo['estado'] or '').lower() != 'pendiente':
        conn.close()
        flash('Solo se pueden editar préstamos en estado pendiente.', 'warning')
        return redirect(url_for('prestamos.detalle_prestamo', pid=pid))

    if request.method == 'POST':
        try:
            monto = float(request.form.get('monto', 0))
            tasa = float(request.form.get('tasa', 0))
            plazo = int(request.form.get('plazo', 0))
            categoria_id = request.form.get('categoria_id')
            
            db_execute(
                conn,
                "UPDATE prestamos SET monto_solicitado=?, tasa_interes=?, plazo_meses=?, categoria_id=? WHERE id=?",
                [monto, tasa, plazo, categoria_id, pid]
            )
            conn.commit()
            flash('Solicitud de préstamo actualizada.', 'success')
            return redirect(url_for('prestamos.detalle_prestamo', pid=pid))
        except Exception as e:
            flash(f'Error al actualizar: {e}', 'danger')
        finally:
            conn.close()
    else:
        categorias = db_fetchall(conn, "SELECT id, nombre FROM prestamo_categorias WHERE estado='activo'")
        configuraciones = db_fetchall(conn, "SELECT * FROM configuraciones WHERE tipo='prestamo'")
        conn.close()
        return render_template('editar_prestamo.html', prestamo=prestamo, categorias=categorias, configuraciones=configuraciones)

@bp.route('/prestamos/eliminar/<int:pid>', methods=['POST'])
@login_required(role='Administrador')
@permission_required('prestamos')
def eliminar_prestamo(pid):
    conn = get_db()
    prestamo = db_fetchone(conn, "SELECT numero, estado FROM prestamos WHERE id=?", [pid])
    
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))
        
    if (prestamo['estado'] or '').lower() != 'pendiente':
        conn.close()
        flash('No se puede eliminar un préstamo que ya ha sido procesado.', 'warning')
        return redirect(url_for('prestamos.detalle_prestamo', pid=pid))

    try:
        db_execute(conn, "DELETE FROM prestamo_calendario_pagos WHERE prestamo_id=?", [pid])
        db_execute(conn, "DELETE FROM prestamos WHERE id=?", [pid])
        conn.commit()
        
        log_auditoria_evento(
            modulo='prestamos',
            entidad='prestamo',
            entidad_id=pid,
            accion='eliminar',
            descripcion=f'Eliminación de solicitud de préstamo {prestamo["numero"]}'
        )
        flash('La solicitud de préstamo ha sido eliminada permanentemente.', 'success')
    except Exception as e:
        flash(f'Error al eliminar: {e}', 'danger')
    finally:
        conn.close()
        
    return redirect(url_for('prestamos.prestamos'))


@bp.route('/prestamos/<int:pid>/calendario/pdf')
def calendario_prestamo_pdf(pid):
    conn = get_db()
    prestamo = obtener_detalle_prestamo_aprobacion(conn, pid)
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    if not prestamo.get('calendario'):
        conn.close()
        flash('Debe generar el calendario de pagos antes de exportarlo.', 'warning')
        return redirect(url_for('prestamos.aprobar_prestamo', pid=pid))

    prestamo['cooperativa_nombre'] = get_system_setting(conn, 'cooperativa_nombre', DEFAULT_COOPERATIVA_NOMBRE)
    conn.close()

    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        flash('ReportLab no está instalado. No se puede generar el PDF. Instálelo con: pip install reportlab', 'danger')
        return redirect(url_for('prestamos.detalle_prestamo', pid=pid))

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 45

    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(40, y, 'Calendario de Pagos de Prestamo')
    y -= 18
    pdf.setFont('Helvetica', 10)
    pdf.drawString(40, y, prestamo['cooperativa_nombre'])
    y -= 16
    pdf.drawString(40, y, f'Prestamo: {prestamo.get("numero") or ""} · Socio: {prestamo.get("socio_codigo") or ""} - {prestamo.get("nombre_socio") or ""}')
    y -= 16
    pdf.drawString(40, y, f'Categoria: {prestamo.get("categoria_nombre") or "General"} · Frecuencia: {prestamo.get("frecuencia") or "Quincenal"}')
    y -= 24

    pdf.setFont('Helvetica-Bold', 9)
    pdf.drawString(40, y, 'Cuota')
    pdf.drawString(110, y, 'Fecha programada')
    pdf.drawString(260, y, 'Monto')
    pdf.drawString(350, y, 'Estado')
    y -= 14
    pdf.line(40, y, width - 40, y)
    y -= 14

    pdf.setFont('Helvetica', 9)
    for cuota in prestamo['calendario']:
        pdf.drawString(40, y, str(cuota['numero_cuota']))
        pdf.drawString(110, y, cuota['fecha_programada'])
        pdf.drawString(260, y, f"Q{float(cuota['monto_programado']):,.2f}")
        pdf.drawString(350, y, cuota.get('estado', 'pendiente').capitalize())
        y -= 16
        if y < 60:
            pdf.showPage()
            y = height - 45
            pdf.setFont('Helvetica', 9)

    pdf.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'calendario_{prestamo.get("numero")}.pdf', mimetype='application/pdf')

@bp.route('/prestamos/<int:pid>/finiquito')
def finiquito_prestamo(pid):
    conn = get_db()
    prestamo = obtener_detalle_prestamo_aprobacion(conn, pid)
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    if prestamo.get('estado') == 'pendiente':
        conn.close()
        flash('Debe aprobar el préstamo antes de generar el finiquito.', 'warning')
        return redirect(url_for('prestamos.aprobar_prestamo', pid=pid))

    prestamo['cooperativa_nombre'] = get_system_setting(conn, 'cooperativa_nombre', DEFAULT_COOPERATIVA_NOMBRE)
    plantilla = get_system_setting(conn, 'prestamo_finiquito_texto', SYSTEM_SETTINGS_DEFAULTS['prestamo_finiquito_texto'])
    contenido = renderizar_finiquito_prestamo(prestamo, plantilla)
    formato = request.args.get('formato', 'html').strip().lower()
    conn.close()

    if formato == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            flash('ReportLab no está instalado. No se puede generar el PDF. Instálelo con: pip install reportlab', 'danger')
            return redirect(url_for('prestamos.finiquito_prestamo', pid=pid, formato='html'))

        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 50

        pdf.setFont('Helvetica-Bold', 14)
        pdf.drawString(40, y, 'Finiquito de Prestamo')
        y -= 22
        pdf.setFont('Helvetica', 10)
        pdf.drawString(40, y, prestamo['cooperativa_nombre'])
        y -= 18
        pdf.drawString(40, y, f'Prestamo: {prestamo.get("numero") or ""}')
        y -= 24

        pdf.setFont('Helvetica', 10)
        for bloque in contenido.split('\n'):
            linea_actual = ''
            for palabra in bloque.split(' '):
                prueba = (linea_actual + ' ' + palabra).strip()
                if pdf.stringWidth(prueba, 'Helvetica', 10) > (width - 80):
                    pdf.drawString(40, y, linea_actual)
                    y -= 16
                    linea_actual = palabra
                else:
                    linea_actual = prueba
            if linea_actual:
                pdf.drawString(40, y, linea_actual)
                y -= 16
            y -= 6
            if y < 70:
                pdf.showPage()
                y = height - 50
                pdf.setFont('Helvetica', 10)

        pdf.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'finiquito_{prestamo.get("numero")}.pdf', mimetype='application/pdf')

    return render_template('finiquito_prestamo.html', prestamo=prestamo, contenido_finiquito=contenido)

@bp.route('/prestamos/<int:pid>/pago', methods=['GET', 'POST'])
@login_required(role=('Administrador', 'Operador'))
@permission_required('prestamos.pagar')
def pagar_prestamo(pid):
    conn = get_db()
    prestamo = db_fetchone(
        conn,
        '''
        SELECT p.*, s.codigo as socio_codigo, s.nombre || ' ' || s.apellido as nombre_socio, s.frecuencia 
        FROM prestamos p 
        JOIN socios s ON s.id = p.socio_id 
        WHERE p.id=?
        ''',
        [pid],
    )
    
    if not prestamo:
        conn.close()
        flash('Préstamo no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    if request.method == 'POST':
        fecha_pago = request.form.get('fecha_pago') or date.today().isoformat()
        # Validar periodo cerrado solamente
        if periodo_cerrado('prestamos', fecha_pago):
            conn.close()
            flash('El periodo de préstamos está cerrado para la fecha seleccionada.', 'warning')
            return redirect(url_for('prestamos.prestamos'))

        
        monto_pago = float(request.form.get('monto', prestamo['cuota_mensual']))
        boleta = request.form.get('boleta_deposito', '').strip()
        fecha_boleta = request.form.get('fecha_boleta') or fecha_pago
        
        tasa_periodica = (prestamo['tasa_interes'] / 100) * (obtener_dias_frecuencia(prestamo['frecuencia']) / 365)
        interes = round(prestamo['saldo_pendiente'] * tasa_periodica, 2)
        capital = round(monto_pago - interes, 2)
        
        if capital <= 0:
            capital = 0
            interes = monto_pago
            
        nuevo_saldo = round(max(0, prestamo['saldo_pendiente'] - capital), 2)
        numero_comprobante = generar_numero_comprobante(conn)
        
        try:
            pago_id = db_insert_and_get_id(
                conn,
                """
                INSERT INTO pagos_prestamo
                (prestamo_id,monto,capital,interes,saldo_restante,fecha,numero_comprobante,boleta_deposito,fecha_boleta,descripcion)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                """,
                [pid, monto_pago, capital, interes, nuevo_saldo, fecha_pago, numero_comprobante, boleta, fecha_boleta, "Cobro de cuota"]
            )

            
            estado = 'pagado' if nuevo_saldo == 0 else 'aprobado'
            db_execute(conn, "UPDATE prestamos SET saldo_pendiente=?, estado=? WHERE id=?", [nuevo_saldo, estado, pid])
            
            cuota_programada = db_fetchone(
                conn,
                '''
                SELECT id FROM prestamo_calendario_pagos
                WHERE prestamo_id=? AND estado='pendiente'
                ORDER BY numero_cuota
                LIMIT 1
                ''',
                [pid]
            )
            if cuota_programada:
                db_execute(
                    conn,
                    "UPDATE prestamo_calendario_pagos SET estado='pagado' WHERE id=?",
                    [cuota_programada['id']]
                )
            
            conn.commit()
            flash(f'Pago registrado exitosamente. Comprobante: {numero_comprobante}', 'success')
            return redirect(url_for('prestamos.detalle_prestamo', pid=pid))
        except Exception as e:
            flash(f'Error al registrar el pago: {e}', 'danger')
        finally:
            if conn: conn.close()
    
    return render_template('registrar_pago.html', prestamo=prestamo, fecha_hoy=date.today().isoformat())



@bp.route('/api/cuota')
def calcular_cuota():
    monto = float(request.args.get('monto', 0))
    tasa = float(request.args.get('tasa', 18))
    plazo = int(request.args.get('plazo', 12))
    tm = tasa / 100 / 12
    cuota = monto * tm / (1 - (1 + tm)**(-plazo)) if tm > 0 else monto / plazo
    return jsonify({'cuota': round(cuota, 2), 'total': round(cuota * plazo, 2), 'intereses': round(cuota * plazo - monto, 2)})

@bp.route('/transacciones_masivas')
@login_required()
def transacciones_masivas():
    return render_template('transacciones_masivas.html')

@bp.route('/menu_prestamos')
@login_required()
def menu_prestamos():
    return render_template('menu_prestamos.html')

@bp.route('/gestiones')
@login_required(role=('Administrador', 'Operador'))
def gestiones():
    conn = get_db()
    tipo_filtro = (request.args.get('tipo') or 'todos').strip().lower()
    destino_filtro = (request.args.get('destino') or 'todos').strip().lower()
    categoria_id_filtro = (request.args.get('categoria_id') or '').strip()
    estados_validos = {'pendiente', 'aprobado', 'no_procede', 'pagado'}
    estado_filtro = (request.args.get('estado') or 'pendiente').strip().lower()
    if estado_filtro not in estados_validos:
        estado_filtro = 'pendiente'
    if destino_filtro not in {'todos', 'retiro', 'amortizacion'}:
        destino_filtro = 'todos'

    categorias_prestamo = db_fetchall(
        conn,
        "SELECT id, nombre FROM prestamo_categorias WHERE estado='activo' ORDER BY nombre"
    )

    solicitudes = []

    if tipo_filtro in ('todos', 'retiro'):
        filtros_retiro = []
        params_retiro = []
        if estado_filtro:
            filtros_retiro.append("sr.estado = ?")
            params_retiro.append(estado_filtro)
        if destino_filtro == 'amortizacion':
            filtros_retiro.append("COALESCE(sr.destino, 'retiro') = 'amortizacion_prestamo'")
        elif destino_filtro == 'retiro':
            filtros_retiro.append("COALESCE(sr.destino, 'retiro') <> 'amortizacion_prestamo'")
        where_retiro = f"WHERE {' AND '.join(filtros_retiro)}" if filtros_retiro else ''
        retiros = db_fetchall(
            conn,
            f'''
            SELECT sr.id,
                   sr.numero,
                   sr.fecha_solicitud,
                   sr.estado,
                   sr.monto,
                   sr.descripcion,
                                     sr.metodo_retiro,
                                     sr.banco_tipo_cuenta,
                                     sr.banco_numero_cuenta,
                                     COALESCE(sr.destino, 'retiro') AS destino,
                                     sr.prestamo_id,
                                     p.numero AS prestamo_numero,
                   s.codigo AS socio_codigo,
                   s.nombre || ' ' || s.apellido AS socio_nombre,
                   c.numero AS cuenta_numero
            FROM solicitudes_retiro sr
            JOIN socios s ON s.id = sr.socio_id
            JOIN cuentas c ON c.id = sr.cuenta_id
                        LEFT JOIN prestamos p ON p.id = sr.prestamo_id
            {where_retiro}
            ORDER BY sr.fecha_solicitud DESC, sr.id DESC
            ''',
            params_retiro,
        )
        for item in retiros:
            row = dict(item)
            row['tipo_solicitud'] = 'retiro'
            row['categoria_prestamo'] = ''
            solicitudes.append(row)

    if tipo_filtro in ('todos', 'prestamo'):
        filtros_prestamo = []
        params_prestamo = []
        if estado_filtro:
            filtros_prestamo.append("p.estado = ?")
            params_prestamo.append(estado_filtro)
        if categoria_id_filtro.isdigit():
            filtros_prestamo.append("p.categoria_id = ?")
            params_prestamo.append(int(categoria_id_filtro))
        where_prestamo = f"WHERE {' AND '.join(filtros_prestamo)}" if filtros_prestamo else ''
        prestamos = db_fetchall(
            conn,
            f'''
            SELECT p.id,
                   p.numero,
                   p.fecha_solicitud,
                   p.estado,
                   p.monto_solicitado AS monto,
                   '' AS descripcion,
                                     '' AS metodo_retiro,
                                     '' AS banco_tipo_cuenta,
                                     '' AS banco_numero_cuenta,
                                     '' AS destino,
                                     NULL AS prestamo_id,
                                     '' AS prestamo_numero,
                   s.codigo AS socio_codigo,
                   s.nombre || ' ' || s.apellido AS socio_nombre,
                   '' AS cuenta_numero,
                   COALESCE(pc.nombre, 'General') AS categoria_prestamo
            FROM prestamos p
            JOIN socios s ON s.id = p.socio_id
            LEFT JOIN prestamo_categorias pc ON pc.id = p.categoria_id
            {where_prestamo}
            ORDER BY p.fecha_solicitud DESC, p.id DESC
            ''',
            params_prestamo,
        )
        for item in prestamos:
            row = dict(item)
            row['tipo_solicitud'] = 'prestamo'
            solicitudes.append(row)

    solicitudes = sorted(
        solicitudes,
        key=lambda item: ((item.get('fecha_solicitud') or ''), item.get('id') or 0),
        reverse=True,
    )

    conn.close()
    return render_template(
        'gestiones.html',
        solicitudes=solicitudes,
        categorias_prestamo=categorias_prestamo,
        tipo_filtro=tipo_filtro,
        destino_filtro=destino_filtro,
        categoria_id_filtro=categoria_id_filtro,
        estado_filtro=estado_filtro,
    )

@bp.route('/gestiones/solicitud-prestamo')
@login_required(role=('Administrador', 'Operador'))
def gestion_solicitud_prestamo():
    return redirect(url_for('prestamos.nuevo_prestamo'))

@bp.route('/planilla_amortizaciones')
@login_required()
def planilla_amortizaciones():
    return render_template('planilla_amortizaciones.html')

@bp.route('/planilla_refinanciamientos')
@login_required()
def planilla_refinanciamientos():
    return render_template('planilla_refinanciamientos.html')

@bp.route('/reportes_prestamos')
@login_required()
def reportes_prestamos():
    return render_template(
        'reportes_prestamos.html',
        fecha_actual=date.today().isoformat(),
        fecha_mes_anterior=(date.today() - timedelta(days=30)).isoformat()
    )

@bp.route('/configuracion_prestamos', methods=['GET', 'POST'])
@login_required()
@permission_required('config.prestamos')
def configuracion_prestamos():
    conn = get_db()
    try:
        ensure_system_settings(conn)
        ensure_module_settings(conn)

        if request.method == 'POST':
            campos = list(PRESTAMO_SETTINGS_DEFAULTS.keys())
            actualizados = 0

            for clave in campos:
                if clave not in request.form:
                    continue
                valor = (request.form.get(clave) or '').strip()
                if not valor:
                    continue
                set_system_setting(conn, clave, valor, session.get('username'))
                actualizados += 1

            conn.commit()
            if actualizados:
                flash('Configuracion de prestamos actualizada correctamente.', 'success')
            else:
                flash('No se recibieron cambios para guardar.', 'warning')
            return redirect(url_for('prestamos.configuracion_prestamos'))

        prestamo_cfg = {
            clave: get_system_setting(conn, clave, valor_default)
            for clave, valor_default in PRESTAMO_SETTINGS_DEFAULTS.items()
        }
        return render_template('configuracion_prestamos.html', prestamo_cfg=prestamo_cfg)
    except Exception as e:
        flash(f'Error cargando configuracion de prestamos: {e}', 'danger')
        return render_template('configuracion_prestamos.html', prestamo_cfg=PRESTAMO_SETTINGS_DEFAULTS)
    finally:
        conn.close()

@bp.route('/cobranza_prestamos')
@login_required()
def cobranza_prestamos():
    return render_template('cobranza_prestamos.html', fecha_actual_datetime=datetime.now().strftime('%Y-%m-%dT%H:%M'))

def _generar_datos_reporte_prestamos(tipo_reporte, fecha_inicio, fecha_fin):
    from utils.prestamos import obtener_cartera_con_alertas
    from utils.helpers import formatear_fecha_dmy
    from datetime import date, datetime, timedelta

    cartera = obtener_cartera_con_alertas(fecha_inicio, fecha_fin)
    activos = [p for p in cartera if p['estado'] == 'aprobado' and p['saldo_pendiente'] > 0]

    total_prestamos = len(activos)
    cartera_total = sum(float(p['saldo_pendiente'] or 0) for p in activos)
    promedio_prestamo = cartera_total / total_prestamos if total_prestamos > 0 else 0.0
    prestamos_vencidos = sum(1 for p in activos if p.get('semaforo') == 'vencido')
    tasa_morosidad = (sum(float(p['saldo_pendiente'] or 0) for p in activos if p.get('semaforo') == 'vencido') / cartera_total * 100) if cartera_total > 0 else 0.0

    conn = get_db()
    intereses_query = "SELECT COALESCE(SUM(interes), 0) FROM pagos_prestamo WHERE 1=1"
    params = []
    if fecha_inicio:
        intereses_query += " AND date(fecha) >= date(?)"
        params.append(fecha_inicio)
    if fecha_fin:
        intereses_query += " AND date(fecha) <= date(?)"
        params.append(fecha_fin)
    intereses_cobrados = db_fetchone(conn, intereses_query, params)[0]
    conn.close()

    rendimiento_cartera = (intereses_cobrados / (cartera_total or 1) * 100)

    estadisticas = {
        'total_prestamos': total_prestamos,
        'cartera_total': round(cartera_total, 2),
        'promedio_prestamo': round(promedio_prestamo, 2),
        'prestamos_vencidos': prestamos_vencidos,
        'tasa_morosidad': round(tasa_morosidad, 1),
        'rendimiento_cartera': round(rendimiento_cartera, 1),
    }

    al_dia = sum(1 for p in activos if p.get('semaforo') == 'al_dia')
    atraso_1_30 = sum(1 for p in activos if p.get('semaforo') == 'vencido' and p.get('dias_atraso', 0) <= 30)
    atraso_31_mas = sum(1 for p in activos if p.get('semaforo') == 'vencido' and p.get('dias_atraso', 0) > 30)

    morosidad = {
        'al_dia': al_dia,
        'atraso_1_30': atraso_1_30,
        'atraso_31_mas': atraso_31_mas
    }

    resultados = []

    if tipo_reporte == 'cartera_activa':
        for p in activos:
            proximo = p.get('proximo_pago')
            if proximo:
                proximo = formatear_fecha_dmy(proximo)
            else:
                proximo = 'N/A'
            resultados.append({
                'numero_prestamo': p['numero'],
                'nombre_socio': p['nombre_socio'],
                'monto_original': float(p['monto_aprobado'] or p['monto_solicitado'] or 0),
                'saldo_actual': float(p['saldo_pendiente'] or 0),
                'cuotas_pendientes': int(p.get('cuotas_pendientes') or 0),
                'proximo_pago': proximo,
                'estado': 'activo'
            })

    elif tipo_reporte == 'morosidad':
        morosos = [p for p in cartera if p['estado'] == 'aprobado' and p['dias_atraso'] > 0]
        for p in morosos:
            ultimo = p.get('ultimo_pago')
            if ultimo:
                ultimo = formatear_fecha_dmy(ultimo)
            else:
                ultimo = 'Ninguno'
            resultados.append({
                'numero_prestamo': p['numero'],
                'nombre_socio': p['nombre_socio'],
                'dias_atraso': int(p['dias_atraso']),
                'monto_vencido': float(p['monto_vencido'] or 0),
                'ultimo_pago': ultimo,
                'estado': p['estado']
            })

    elif tipo_reporte == 'pagos_vencidos':
        conn = get_db()
        rows = db_fetchall(conn, '''
            SELECT cp.fecha_programada, cp.monto_programado, p.numero AS numero_prestamo,
                   s.nombre || ' ' || s.apellido AS nombre_socio
            FROM prestamo_calendario_pagos cp
            JOIN prestamos p ON cp.prestamo_id = p.id
            JOIN socios s ON p.socio_id = s.id
            WHERE cp.estado = 'pendiente' AND p.estado = 'aprobado'
        ''')
        conn.close()
        hoy = date.today()
        for r in rows:
            f_prog = date.fromisoformat(r['fecha_programada'])
            if f_prog < hoy:
                dias = (hoy - f_prog).days
                resultados.append({
                    'numero_prestamo': r['numero_prestamo'],
                    'nombre_socio': r['nombre_socio'],
                    'fecha_vencimiento': formatear_fecha_dmy(r['fecha_programada']),
                    'monto_vencido': float(r['monto_programado'] or 0),
                    'dias_atraso': dias,
                    'estado': 'Vencido'
                })

    elif tipo_reporte == 'rendimiento':
        conn = get_db()
        pagos = db_fetchall(conn, "SELECT fecha, interes, capital, monto FROM pagos_prestamo")
        conn.close()
        por_mes = {}
        for p in pagos:
            mes = p['fecha'][:7]
            por_mes.setdefault(mes, {'interes': 0.0, 'capital': 0.0, 'monto': 0.0})
            por_mes[mes]['interes'] += float(p['interes'] or 0)
            por_mes[mes]['capital'] += float(p['capital'] or 0)
            por_mes[mes]['monto'] += float(p['monto'] or 0)
            
        for mes in sorted(por_mes.keys(), reverse=True):
            y, m = mes.split('-')
            mes_fmt = f"{m}/{y}"
            cartera_prom = float(cartera_total or 100000)
            int_cob = por_mes[mes]['interes']
            rend = (int_cob / cartera_prom * 100)
            resultados.append({
                'mes': mes_fmt,
                'intereses_cobrados': int_cob,
                'morosidad': float(tasa_morosidad),
                'cartera_promedio': cartera_prom,
                'rendimiento': rend
            })

    elif tipo_reporte == 'comparativo':
        conn = get_db()
        prestamos_aprobados = db_fetchall(conn, "SELECT fecha_aprobacion, monto_aprobado FROM prestamos WHERE estado IN ('aprobado', 'pagado') AND fecha_aprobacion IS NOT NULL")
        conn.close()
        nuevos_por_mes = {}
        for p in prestamos_aprobados:
            mes = p['fecha_aprobacion'][:7]
            nuevos_por_mes.setdefault(mes, {'count': 0, 'monto': 0.0})
            nuevos_por_mes[mes]['count'] += 1
            nuevos_por_mes[mes]['monto'] += float(p['monto_aprobado'] or 0)
            
        hoy = date.today()
        for i in range(6):
            d = hoy - timedelta(days=i*30)
            mes = d.isoformat()[:7]
            y, m = mes.split('-')
            mes_fmt = f"{m}/{y}"
            nuevos = nuevos_por_mes.get(mes, {'count': 0, 'monto': 0.0})
            c_actual = max(10000, float(cartera_total) - i * 5000)
            c_anterior = max(10000, c_actual - nuevos['monto'] + 2000)
            resultados.append({
                'mes': mes_fmt,
                'nuevos_prestamos': nuevos['count'],
                'cartera_anterior': c_anterior,
                'cartera_actual': c_actual
            })

    elif tipo_reporte == 'riesgo':
        for p in activos:
            dias = int(p.get('dias_atraso') or 0)
            if dias == 0:
                score = 95
                historial = 'excelente'
                capacidad = 'alta'
                nivel = 'bajo'
            elif dias <= 30:
                score = 75
                historial = 'bueno'
                capacidad = 'media'
                nivel = 'medio'
            else:
                score = 45
                historial = 'regular'
                capacidad = 'baja'
                nivel = 'alto'
            resultados.append({
                'numero_prestamo': p['numero'],
                'nombre_socio': p['nombre_socio'],
                'score_riesgo': score,
                'historial_pagos': historial,
                'capacidad_pago': capacidad,
                'nivel_riesgo': nivel
            })

    return resultados, estadisticas, morosidad


@bp.route('/generar_reporte_prestamos', methods=['POST'])
@login_required()
def generar_reporte_prestamos():
    data = request.get_json() or {}
    tipo_reporte = (data.get('tipo_reporte') or 'cartera_activa').strip()
    fecha_inicio = (data.get('fecha_inicio') or '').strip() or None
    fecha_fin = (data.get('fecha_fin') or '').strip() or None

    try:
        resultados, estadisticas, morosidad = _generar_datos_reporte_prestamos(tipo_reporte, fecha_inicio, fecha_fin)
        return jsonify({'success': True, 'resultados': resultados, 'estadisticas': estadisticas, 'morosidad': morosidad})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@bp.route('/reporte_prestamos/export')
@login_required()
def exportar_reporte_prestamos():
    tipo_reporte = request.args.get('tipo_reporte', 'cartera_activa').strip()
    fecha_inicio = request.args.get('fecha_inicio', '').strip() or None
    fecha_fin = request.args.get('fecha_fin', '').strip() or None
    formato = request.args.get('formato', 'excel').strip().lower()
    resultados, _, _ = _generar_datos_reporte_prestamos(tipo_reporte, fecha_inicio, fecha_fin)

    if formato == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            flash('ReportLab no está instalado. No se puede generar el PDF. Instálelo con: pip install reportlab', 'danger')
            return redirect(url_for('prestamos.reportes_prestamos'))

        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 40

        pdf.setFont('Helvetica-Bold', 12)
        pdf.drawString(40, y, f'Reporte de Prestamos: {tipo_reporte}')
        y -= 16
        pdf.setFont('Helvetica', 9)
        pdf.drawString(40, y, f'Fecha de exportacion: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
        y -= 24

        if resultados:
            columnas = list(resultados[0].keys())
            pdf.setFont('Helvetica-Bold', 8)
            x = 40
            for col in columnas[:6]:
                pdf.drawString(x, y, str(col)[:18])
                x += 88
            y -= 14
            pdf.setFont('Helvetica', 7)

            for row in resultados:
                x = 40
                for col in columnas[:6]:
                    valor = str(row.get(col, ''))
                    pdf.drawString(x, y, valor[:18])
                    x += 88
                y -= 12
                if y < 60:
                    pdf.showPage()
                    y = height - 40
                    pdf.setFont('Helvetica', 7)
        else:
            pdf.setFont('Helvetica', 10)
            pdf.drawString(40, y, 'Sin datos para el rango seleccionado.')

        pdf.save()
        buffer.seek(0)
        filename = f"reporte_prestamos_{tipo_reporte}_{date.today().isoformat()}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

    if formato == 'csv':
        output = StringIO()
        if resultados:
            writer = csv.DictWriter(output, fieldnames=list(resultados[0].keys()))
            writer.writeheader()
            writer.writerows(resultados)
        else:
            output.write('sin_datos\n')
        filename = f"reporte_prestamos_{tipo_reporte}_{date.today().isoformat()}.csv"
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Prestamos'
        if resultados:
            headers = list(resultados[0].keys())
            ws.append(headers)
            for row in resultados:
                ws.append([row.get(h) for h in headers])
        else:
            ws.append(['Sin datos'])
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        filename = f"reporte_prestamos_{tipo_reporte}_{date.today().isoformat()}.xlsx"
        return send_file(
            file_data,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except ImportError:
        flash('La exportación a Excel no está disponible (falta instalar openpyxl).', 'warning')
        return redirect(url_for('prestamos.reportes_prestamos'))

@bp.route('/obtener_estadisticas_cobranza')
@login_required()
def obtener_estadisticas_cobranza():
    cartera = obtener_cartera_con_alertas()
    morosos = [p for p in cartera if p['dias_atraso'] > 0 and p['estado'] == 'aprobado']
    conn = get_db()
    if _is_postgres_connection(conn):
        recuperado_mes = db_fetchone(
            conn,
            """
            SELECT COALESCE(SUM(monto),0)
            FROM pagos_prestamo
            WHERE to_char(fecha::date, 'YYYY-MM') = to_char(current_date, 'YYYY-MM')
            """
        )[0]
        acciones_pendientes = db_fetchone(
            conn,
            """
            SELECT COUNT(*)
            FROM cobranza_acciones
            WHERE fecha_compromiso::date < current_date
              AND resultado IN ('compromiso', 'sin_respuesta')
            """
        )[0]
    else:
        recuperado_mes = db_fetchone(
            conn,
            """
            SELECT COALESCE(SUM(monto),0)
            FROM pagos_prestamo
            WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            """
        )[0]
        acciones_pendientes = db_fetchone(
            conn,
            """
            SELECT COUNT(*)
            FROM cobranza_acciones
            WHERE date(fecha_compromiso) < date('now')
              AND resultado IN ('compromiso', 'sin_respuesta')
            """
        )[0]
    conn.close()
    return jsonify({
        'prestamos_morosos': len(morosos),
        'monto_moroso': float(sum(p['monto_vencido'] for p in morosos)),
        'acciones_pendientes': int(acciones_pendientes),
        'recuperado_mes': float(recuperado_mes or 0),
    })

@bp.route('/obtener_lista_cobranza', methods=['POST'])
@login_required()
def obtener_lista_cobranza():
    data = request.get_json() or {}
    filtro = (data.get('filtro_morosidad') or 'todos').strip()
    ordenar_por = (data.get('ordenar_por') or 'dias_atraso').strip()
    filtro_socio = (data.get('filtro_socio') or '').strip().lower()
    filtro_responsable = (data.get('filtro_responsable') or '').strip().lower()

    cartera = [p for p in obtener_cartera_con_alertas() if p['estado'] == 'aprobado' and p['dias_atraso'] > 0]

    if filtro == '1-30':
        cartera = [p for p in cartera if 1 <= p['dias_atraso'] <= 30]
    elif filtro == '31-60':
        cartera = [p for p in cartera if 31 <= p['dias_atraso'] <= 60]
    elif filtro == '61-90':
        cartera = [p for p in cartera if 61 <= p['dias_atraso'] <= 90]
    elif filtro == '90+':
        cartera = [p for p in cartera if p['dias_atraso'] > 90]

    key_map = {
        'dias_atraso': lambda x: x['dias_atraso'],
        'monto_vencido': lambda x: x['monto_vencido'],
        'fecha_ultimo_pago': lambda x: x.get('ultimo_pago') or '',
        'numero_prestamo': lambda x: x['numero'],
    }
    cartera.sort(key=key_map.get(ordenar_por, key_map['dias_atraso']), reverse=(ordenar_por != 'numero_prestamo'))

    conn = get_db()
    ult_contactos = db_fetchall(
        conn,
        '''
        SELECT p.numero AS numero_prestamo, MAX(ca.fecha_accion) AS ultimo_contacto
        FROM cobranza_acciones ca
        JOIN prestamos p ON ca.prestamo_id = p.id
        GROUP BY p.numero
        '''
    )
    ult_responsables = db_fetchall(
        conn,
        '''
        SELECT p.numero AS numero_prestamo, ca.responsable
        FROM cobranza_acciones ca
        JOIN prestamos p ON ca.prestamo_id = p.id
        JOIN (
            SELECT prestamo_id, MAX(id) AS ultimo_id
            FROM cobranza_acciones
            GROUP BY prestamo_id
        ) ult ON ult.ultimo_id = ca.id
        '''
    )
    conn.close()
    mapa_contacto = {r['numero_prestamo']: r['ultimo_contacto'] for r in ult_contactos}
    mapa_responsable = {r['numero_prestamo']: (r['responsable'] or '') for r in ult_responsables}

    respuesta = []
    for p in cartera:
        etapa = (p.get('etapa_cobranza') or 'activo').lower()
        responsable = mapa_responsable.get(p['numero'], '')

        if filtro_socio:
            texto_socio = f"{(p.get('socio_codigo') or '').lower()} {(p.get('nombre_socio') or '').lower()}"
            if filtro_socio not in texto_socio:
                continue

        if filtro_responsable and filtro_responsable not in responsable.lower():
            continue

        respuesta.append({
            'numero_prestamo': p['numero'],
            'nombre_socio': p['nombre_socio'],
            'dias_atraso': int(p['dias_atraso']),
            'monto_vencido': float(p['monto_vencido']),
            'ultimo_pago': p['ultimo_pago'],
            'ultimo_contacto': mapa_contacto.get(p['numero']),
            'responsable': responsable or 'Sin asignar',
            'estado_cobranza': etapa,
        })

    return jsonify({'prestamos': respuesta})

@bp.route('/guardar_accion_cobranza', methods=['POST'])
@login_required()
@permission_required('cobranza.gestion')
def guardar_accion_cobranza():
    data = request.get_json() or {}
    numero_prestamo = (data.get('numero_prestamo') or '').strip()
    if not numero_prestamo:
        return jsonify({'success': False, 'error': 'Número de préstamo requerido'}), 400

    conn = get_db()
    prestamo = db_fetchone(conn, 'SELECT id, numero FROM prestamos WHERE numero=?', (numero_prestamo,))
    if not prestamo:
        conn.close()
        return jsonify({'success': False, 'error': 'Préstamo no encontrado'}), 404

    db_execute(
        conn,
        '''
        INSERT INTO cobranza_acciones
        (prestamo_id, tipo_accion, resultado, notas, monto_comprometido, fecha_compromiso, fecha_accion, responsable)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            prestamo['id'],
            data.get('tipo_accion', 'llamada'),
            data.get('resultado', 'sin_respuesta'),
            data.get('notas', ''),
            float(data.get('monto_comprometido', 0) or 0),
            data.get('fecha_compromiso') or None,
            data.get('fecha_accion') or datetime.now().isoformat(),
            session.get('username', 'operador'),
        ),
    )
    conn.commit()
    conn.close()

    log_auditoria_evento(
        modulo='prestamos',
        entidad='cobranza_accion',
        entidad_id=prestamo['id'],
        accion='crear',
        descripcion=f'Acción de cobranza para préstamo {numero_prestamo}',
        datos=data
    )

    return jsonify({'success': True})

@bp.route('/obtener_historial_cobranza')
@login_required()
def obtener_historial_cobranza():
    responsable = request.args.get('responsable', '').strip().lower()
    numero = request.args.get('numero_prestamo', '').strip().lower()

    filtros = 'WHERE 1=1'
    params = []
    if responsable:
        filtros += ' AND lower(COALESCE(ca.responsable, "")) LIKE ?'
        params.append(f'%{responsable}%')
    if numero:
        filtros += ' AND lower(p.numero) LIKE ?'
        params.append(f'%{numero}%')

    conn = get_db()
    rows = db_fetchall(
        conn,
        f'''
        SELECT ca.fecha_accion AS fecha,
               p.numero AS numero_prestamo,
               ca.tipo_accion,
               ca.resultado,
               ca.responsable
        FROM cobranza_acciones ca
        JOIN prestamos p ON p.id = ca.prestamo_id
        {filtros}
        ORDER BY ca.id DESC
        LIMIT 200
        ''',
        params
    )
    conn.close()
    return jsonify({'historial': [dict(r) for r in rows]})

@bp.route('/enviar_recordatorios_cobranza', methods=['POST'])
@login_required()
@permission_required('cobranza.recordatorios')
def enviar_recordatorios_cobranza():
    data = request.get_json() or {}
    numeros = data.get('prestamos', [])
    if not numeros:
        return jsonify({'message': 'No se seleccionaron préstamos.'}), 400

    conn = get_db()
    enviados = 0
    for numero in numeros:
        prestamo = db_fetchone(conn, 'SELECT id FROM prestamos WHERE numero=?', (numero,))
        if not prestamo:
            continue
        db_execute(
            conn,
            '''
            INSERT INTO cobranza_acciones
            (prestamo_id, tipo_accion, resultado, notas, fecha_accion, responsable)
            VALUES (?, 'recordatorio', 'contactado', 'Recordatorio automático generado desde panel', ?, ?)
            ''',
            (prestamo['id'], datetime.now().isoformat(), session.get('username', 'operador')),
        )
        enviados += 1
    conn.commit()
    conn.close()
    return jsonify({'message': f'Recordatorios generados para {enviados} préstamos.'})

@bp.route('/marcar_revision_legal', methods=['POST'])
@login_required()
@permission_required('cobranza.legal')
def marcar_revision_legal():
    data = request.get_json() or {}
    numeros = data.get('prestamos', [])
    if not numeros:
        return jsonify({'message': 'No se seleccionaron préstamos.'}), 400

    conn = get_db()
    marcados = 0
    for numero in numeros:
        db_execute(conn, "UPDATE prestamos SET etapa_cobranza='legal' WHERE numero=?", (numero,))
        cur_count = db_fetchone(conn, "SELECT COUNT(*) FROM prestamos WHERE numero=? AND etapa_cobranza='legal'", (numero,))[0]
        if cur_count:
            marcados += 1
    conn.commit()
    conn.close()

    log_auditoria_evento(
        modulo='prestamos',
        entidad='prestamo',
        accion='marcar_legal',
        descripcion='Préstamos marcados para revisión legal',
        datos={'prestamos': numeros, 'total': marcados}
    )

    return jsonify({'message': f'{marcados} préstamos enviados a revisión legal.'})

@bp.route('/socios/<int:sid>/estado_cuenta_prestamo')
@login_required()
def estado_cuenta_prestamo(sid):
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    export = request.args.get('export', '').strip().lower()
    printable = request.args.get('print', '').strip() == '1'

    conn = get_db()
    socio = db_fetchone(conn, 'SELECT * FROM socios WHERE id=?', (sid,))
    if not socio:
        conn.close()
        flash('Socio no encontrado.', 'danger')
        return redirect(url_for('socios.socios'))

    filtros = ''
    params = [sid]
    if fecha_desde:
        filtros += ' AND date(pp.fecha) >= date(?)'
        params.append(fecha_desde)
    if fecha_hasta:
        filtros += ' AND date(pp.fecha) <= date(?)'
        params.append(fecha_hasta)

    pagos = db_fetchall(
        conn,
        f'''
        SELECT pp.*, p.numero AS numero_prestamo
        FROM pagos_prestamo pp
        JOIN prestamos p ON p.id = pp.prestamo_id
        WHERE p.socio_id=? {filtros}
        ORDER BY date(pp.fecha) DESC, pp.id DESC
        ''',
        params,
    )

    resumen = db_fetchone(
        conn,
        '''
        SELECT COUNT(*) AS total_prestamos,
               COALESCE(SUM(CASE WHEN estado='aprobado' THEN saldo_pendiente ELSE 0 END),0) AS saldo_activo,
               COALESCE(SUM(CASE WHEN estado='pagado' THEN 1 ELSE 0 END),0) AS prestamos_cancelados
        FROM prestamos
        WHERE socio_id=?
        ''',
        (sid,),
    )
    conn.close()

    total_pagado = sum(float(p['monto'] or 0) for p in pagos)

    if export == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Fecha', 'Prestamo', 'Monto', 'Capital', 'Interes', 'Saldo Restante', 'Comprobante', 'Boleta'])
        for p in pagos:
            writer.writerow([
                p['fecha'], p['numero_prestamo'], p['monto'], p['capital'], p['interes'], p['saldo_restante'],
                p['numero_comprobante'] or '', p['boleta_deposito'] or ''
            ])
        filename = f"estado_cuenta_prestamo_{socio['codigo']}_{date.today().isoformat()}.csv"
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    if export == 'excel':
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Estado Cuenta'
        ws.append(['Fecha', 'Prestamo', 'Monto', 'Capital', 'Interes', 'Saldo Restante', 'Comprobante', 'Boleta'])
        for p in pagos:
            ws.append([
                p['fecha'], p['numero_prestamo'], float(p['monto'] or 0), float(p['capital'] or 0),
                float(p['interes'] or 0), float(p['saldo_restante'] or 0), p['numero_comprobante'] or '', p['boleta_deposito'] or ''
            ])
        mem = BytesIO()
        wb.save(mem)
        mem.seek(0)
        filename = f"estado_cuenta_prestamo_{socio['codigo']}_{date.today().isoformat()}.xlsx"
        return send_file(
            mem,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template(
        'estado_cuenta_prestamo.html',
        socio=socio,
        pagos=pagos,
        resumen=resumen,
        total_pagado=total_pagado,
        filtros={'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta},
        printable=printable,
    )

@bp.route('/prestamos/comprobante/<int:pago_id>')
@login_required()
def comprobante_pago_prestamo(pago_id):
    conn = get_db()
    pago = db_fetchone(
        conn,
        '''
        SELECT pp.*, p.numero AS numero_prestamo,
               s.codigo AS socio_codigo,
               s.nombre || ' ' || s.apellido AS socio_nombre
        FROM pagos_prestamo pp
        JOIN prestamos p ON p.id = pp.prestamo_id
        JOIN socios s ON s.id = p.socio_id
        WHERE pp.id=?
        ''',
        (pago_id,),
    )
    conn.close()

    if not pago:
        flash('Comprobante no encontrado.', 'danger')
        return redirect(url_for('prestamos.prestamos'))

    return render_template('comprobante_prestamo.html', pago=pago)

@bp.route('/planilla_prestamos')
@login_required()
def planilla_prestamos():
    return redirect(url_for('prestamos.planillas_prestamos_pendientes'))

@bp.route('/planillas_prestamos_pendientes')
@login_required()
def planillas_prestamos_pendientes():
    import math as _math
    conn = get_db()
    nombre = request.args.get('nombre', '').strip()
    frecuencia = request.args.get('frecuencia', '').strip()
    estado = request.args.get('estado', '').strip().lower()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = min(100, max(10, int(request.args.get('per_page', 50) or 50)))

    base_query = "FROM planillas_masivas WHERE tipo = 'prestamo_cuotas'"
    params = []

    if nombre:
        base_query += ' AND nombre LIKE ?'
        params.append(f'%{nombre}%')
    if frecuencia:
        base_query += ' AND frecuencia = ?'
        params.append(frecuencia)
    if estado:
        base_query += ' AND estado = ?'
        params.append(estado)
    if fecha_desde:
        base_query += ' AND date(fecha_pago) >= date(?)'
        params.append(fecha_desde)
    if fecha_hasta:
        base_query += ' AND date(fecha_pago) <= date(?)'
        params.append(fecha_hasta)

    order_sql = '''
        ORDER BY CASE estado
            WHEN 'pendiente' THEN 1
            WHEN 'parcial' THEN 2
            WHEN 'aplicada' THEN 3
            ELSE 4
        END, fecha_creacion DESC, id DESC
    '''

    total_planillas = db_fetchone(conn, f'SELECT COUNT(*) {base_query}', params)[0]
    total_pages = max(1, _math.ceil(total_planillas / per_page))
    offset = (page - 1) * per_page

    planillas = db_fetchall(
        conn,
        f'SELECT * {base_query} {order_sql} LIMIT ? OFFSET ?',
        params + [per_page, offset]
    )
    total_monto_row = db_fetchone(conn, f'SELECT COALESCE(SUM(total_monto),0) {base_query}', params)
    total_monto = float(total_monto_row[0] or 0)
    conn.close()

    return render_template(
        'planillas_prestamos_pendientes.html',
        planillas=planillas,
        filtros={
            'nombre': nombre,
            'frecuencia': frecuencia,
            'estado': estado,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta
        },
        total_planillas=total_planillas,
        total_monto=total_monto,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )

@bp.route('/planillas_prestamos/<int:planilla_id>')
@login_required()
def detalle_planilla_prestamos(planilla_id):
    conn = get_db()
    planilla = db_fetchone(conn, '''
        SELECT * FROM planillas_masivas
        WHERE id=? AND tipo='prestamo_cuotas'
    ''', (planilla_id,))

    if not planilla:
        conn.close()
        flash('Planilla de prestamos no encontrada.', 'danger')
        return redirect(url_for('prestamos.planillas_prestamos_pendientes'))

    detalles = db_fetchall(conn, '''
        SELECT d.*, p.monto_aprobado, p.saldo_pendiente, p.cuota_mensual,
               COALESCE(pp.capital_pagado, 0) AS capital_pagado,
               COALESCE(pp.interes_pagado, 0) AS interes_pagado
        FROM planilla_masiva_detalles d
        LEFT JOIN prestamos p ON d.referencia_id = p.id AND d.referencia_tipo = 'prestamo'
        LEFT JOIN (
            SELECT prestamo_id,
                   SUM(capital) AS capital_pagado,
                   SUM(interes) AS interes_pagado
            FROM pagos_prestamo
            GROUP BY prestamo_id
        ) pp ON pp.prestamo_id = p.id
        WHERE d.planilla_id=?
        ORDER BY socio_nombre, numero_referencia
    ''', (planilla_id,))
    conn.close()

    return render_template(
        'planilla_prestamos.html',
        planilla=planilla,
        detalles=detalles,
        nombre_planilla=planilla['nombre'],
        fecha_pago=planilla['fecha_pago'],
        boleta_deposito=planilla['boleta_deposito'],
        frecuencia=planilla['frecuencia']
    )

@bp.route('/planillas_prestamos/<int:planilla_id>/exportar')
@login_required()
def exportar_planilla_prestamos(planilla_id):
    conn = get_db()
    planilla = db_fetchone(conn, '''
        SELECT * FROM planillas_masivas
        WHERE id=? AND tipo='prestamo_cuotas'
    ''', (planilla_id,))

    if not planilla:
        conn.close()
        flash('Planilla de prestamos no encontrada.', 'danger')
        return redirect(url_for('prestamos.planillas_prestamos_pendientes'))

    detalles = db_fetchall(conn, '''
        SELECT d.*, p.monto_aprobado, p.saldo_pendiente, p.cuota_mensual,
               COALESCE(pp.capital_pagado, 0) AS capital_pagado,
               COALESCE(pp.interes_pagado, 0) AS interes_pagado
        FROM planilla_masiva_detalles d
        LEFT JOIN prestamos p ON d.referencia_id = p.id AND d.referencia_tipo = 'prestamo'
        LEFT JOIN (
            SELECT prestamo_id,
                   SUM(capital) AS capital_pagado,
                   SUM(interes) AS interes_pagado
            FROM pagos_prestamo
            GROUP BY prestamo_id
        ) pp ON pp.prestamo_id = p.id
        WHERE d.planilla_id=?
        ORDER BY socio_nombre, numero_referencia
    ''', (planilla_id,))
    conn.close()

    import io
    import csv
    from flask import Response

    output = io.StringIO()
    output.write('\ufeff')  # BOM de UTF-8 para Excel
    writer = csv.writer(output, delimiter=';')

    headers = [
        'Código Socio', 'Nombre Socio', 'No. Préstamo', 
        'Monto Aprobado', 'Saldo Pendiente', 'Cuota Mensual', 
        'Monto a Cobrar', 'Estado'
    ]
    writer.writerow(headers)

    for d in detalles:
        writer.writerow([
            d['socio_codigo'],
            d['socio_nombre'],
            d['numero_referencia'],
            f"Q{d['monto_aprobado']:.2f}" if d['monto_aprobado'] is not None else "Q0.00",
            f"Q{d['saldo_pendiente']:.2f}" if d['saldo_pendiente'] is not None else "Q0.00",
            f"Q{d['cuota_mensual']:.2f}" if d['cuota_mensual'] is not None else "Q0.00",
            f"Q{d['monto']:.2f}",
            d['estado'].upper()
        ])

    filename = f"Planilla_Prestamos_{planilla['nombre'].replace(' ', '_')}"
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
    response.headers['Content-type'] = 'text/csv; charset=utf-8'
    return response

@bp.route('/generar_planilla_prestamos', methods=['GET', 'POST'])
@login_required()
def generar_planilla_prestamos():
    form_data = {
        'nombre_planilla': '',
        'fecha_pago': date.today().isoformat(),
        'frecuencia': 'Quincenal',
    }

    if request.method == 'POST':
        nombre_planilla = request.form.get('nombre_planilla', '').strip()
        fecha_pago = request.form.get('fecha_pago', '').strip()
        frecuencia = request.form.get('frecuencia', '').strip()

        form_data = {
            'nombre_planilla': nombre_planilla,
            'fecha_pago': fecha_pago or date.today().isoformat(),
            'frecuencia': frecuencia or 'Quincenal',
        }

        if not nombre_planilla or not fecha_pago or not frecuencia:
            flash('Todos los campos son obligatorios.', 'danger')
            return render_template('generar_planilla_prestamos.html', form_data=form_data)

        if frecuencia not in ('Quincenal', 'Catorcenal'):
            flash('Frecuencia no valida.', 'danger')
            return render_template('generar_planilla_prestamos.html', form_data=form_data)

        conn = get_db()

        # Obtener prestamos activos filtrados por frecuencia del socio y con cuotas pendientes hasta la fecha seleccionada.
        prestamos = db_fetchall(
            conn,
            '''
            SELECT p.id, p.numero, p.monto_aprobado, p.saldo_pendiente, p.cuota_mensual,
                   p.tasa_interes, p.plazo_meses, s.id AS socio_id, s.nombre, s.apellido,
                   s.codigo, s.frecuencia, COUNT(pp.id) AS cuotas_pagadas
            FROM prestamos p
            JOIN socios s ON p.socio_id = s.id
            LEFT JOIN pagos_prestamo pp ON p.id = pp.prestamo_id
            WHERE p.estado = 'aprobado'
              AND p.saldo_pendiente > 0
              AND s.estado = 'activo'
              AND s.frecuencia = ?
              AND p.id IN (
                  SELECT DISTINCT prestamo_id
                  FROM prestamo_calendario_pagos
                  WHERE estado = 'pendiente' AND fecha_programada <= ?
              )
            GROUP BY p.id
            ORDER BY s.apellido, s.nombre
            ''',
            (frecuencia, fecha_pago)
        )

        if not prestamos:
            conn.close()
            flash('No se encontraron prestamos para generar la planilla con los filtros seleccionados.', 'warning')
            return render_template('generar_planilla_prestamos.html', form_data=form_data)

        total_planilla = sum(min(float(prestamo['cuota_mensual'] or 0), float(prestamo['saldo_pendiente'] or 0)) for prestamo in prestamos)

        planilla_id = db_insert_and_get_id(
            conn,
            '''
            INSERT INTO planillas_masivas
            (tipo, nombre, fecha_pago, frecuencia, estado, total_monto, total_registros, fecha_creacion, usuario_creacion)
            VALUES (?, ?, ?, ?, 'pendiente', ?, ?, ?, ?)
            ''',
            (
                'prestamo_cuotas', nombre_planilla, fecha_pago, frecuencia,
                total_planilla, len(prestamos), datetime.now().isoformat(), session.get('username')
            )
        )

        for prestamo in prestamos:
            monto_programado = min(float(prestamo['cuota_mensual'] or 0), float(prestamo['saldo_pendiente'] or 0))
            db_execute(
                conn,
                '''
                INSERT INTO planilla_masiva_detalles
                (planilla_id, referencia_tipo, referencia_id, numero_referencia, socio_codigo, socio_nombre, monto, estado)
                VALUES (?, 'prestamo', ?, ?, ?, ?, ?, 'pendiente')
                ''',
                (
                    planilla_id,
                    prestamo['id'],
                    prestamo['numero'],
                    prestamo['codigo'],
                    f"{prestamo['nombre']} {prestamo['apellido']}",
                    monto_programado
                )
            )

        conn.commit()
        conn.close()
        flash('Planilla de prestamos generada y guardada como pendiente.', 'success')
        return redirect(url_for('prestamos.detalle_planilla_prestamos', planilla_id=planilla_id))

    return render_template('generar_planilla_prestamos.html', form_data=form_data)

@bp.route('/procesar_pagos_masivos', methods=['POST'])
@login_required()
@permission_required('prestamos.masivo')
def procesar_pagos_masivos():
    data = request.get_json()
    planilla_id = data.get('planilla_id')
    pagos = data.get('pagos', [])
    fecha_pago = data.get('fecha_pago', date.today().isoformat())
    nombre_planilla = data.get('nombre_planilla', 'Planilla de prestamos').strip()
    boleta_deposito = data.get('boleta_deposito', '').strip()
    frecuencia = data.get('frecuencia', '').strip()

    if periodo_cerrado('prestamos', fecha_pago):
        return jsonify({'error': 'El periodo de préstamos está cerrado para la fecha indicada.'}), 400

    if not boleta_deposito:
        return jsonify({'error': 'Debe indicar numero de boleta de pago para aplicar la planilla.'}), 400
    
    conn = get_db()
    planilla = None
    if planilla_id:
        planilla = db_fetchone(conn, '''
            SELECT * FROM planillas_masivas
            WHERE id=? AND tipo='prestamo_cuotas'
        ''', (planilla_id,))

        if not planilla:
            conn.close()
            return jsonify({'error': 'La planilla seleccionada no existe.'}), 404

        if planilla['estado'] == 'aplicada':
            conn.close()
            return jsonify({'error': 'La planilla ya fue aplicada anteriormente.'}), 400
    
    procesados = 0
    errores = []
    total_capital = 0.0
    total_interes = 0.0
    resumen_aplicados = []
    
    for pago in pagos:
        try:
            prestamo_id = pago['prestamo_id']
            monto = float(pago['monto'])
            
            if monto <= 0:
                errores.append(f"Monto inválido para préstamo {pago.get('numero', prestamo_id)}")
                continue
            
            # Obtener información del préstamo
            prestamo = db_fetchone(conn, '''
                SELECT p.saldo_pendiente, p.cuota_mensual, p.socio_id, s.frecuencia, p.tasa_interes
                FROM prestamos p
                JOIN socios s ON p.socio_id = s.id
                WHERE p.id = ? AND p.estado = "aprobado"
            ''', (prestamo_id,))
            
            if not prestamo:
                errores.append(f"Préstamo {pago.get('numero', prestamo_id)} no encontrado o no aprobado")
                continue
            
            saldo_pendiente = float(prestamo['saldo_pendiente'] or 0)
            cuota_mensual = float(prestamo['cuota_mensual'] or 0)

            if frecuencia and prestamo['frecuencia'] != frecuencia:
                errores.append(f"Prestamo {pago.get('numero', prestamo_id)} no coincide con la frecuencia seleccionada")
                continue
            
            if monto > saldo_pendiente:
                errores.append(f"Monto excede saldo pendiente para préstamo {pago.get('numero', prestamo_id)}")
                continue

            # Calcular capital e intereses usando la misma fórmula que pagar_prestamo
            tasa_periodica = (prestamo['tasa_interes'] / 100) * (obtener_dias_frecuencia(prestamo['frecuencia']) / 365)
            interes = round(saldo_pendiente * tasa_periodica, 2)
            capital = round(monto - interes, 2)
            if capital <= 0:
                capital = round(monto, 2)
                interes = 0.0
            
            nuevo_saldo = round(max(0, saldo_pendiente - monto), 2)
            
            # Actualizar saldo del préstamo y su estado si se ha cancelado
            estado_prestamo = 'pagado' if nuevo_saldo <= 0 else 'aprobado'
            db_execute(
                conn,
                'UPDATE prestamos SET saldo_pendiente = ?, estado = ? WHERE id = ?',
                (nuevo_saldo, estado_prestamo, prestamo_id)
            )

            # Actualizar calendario de pagos (marcar cuota(s) como pagada(s))
            if nuevo_saldo <= 0:
                db_execute(
                    conn,
                    "UPDATE prestamo_calendario_pagos SET estado='pagado' WHERE prestamo_id=? AND estado='pendiente'",
                    [prestamo_id]
                )
            else:
                cuota_programada = db_fetchone(
                    conn,
                    '''
                    SELECT id FROM prestamo_calendario_pagos
                    WHERE prestamo_id=? AND estado='pendiente'
                    ORDER BY numero_cuota
                    LIMIT 1
                    ''',
                    [prestamo_id]
                )
                if cuota_programada:
                    db_execute(
                        conn,
                        "UPDATE prestamo_calendario_pagos SET estado='pagado' WHERE id=?",
                        [cuota_programada['id']]
                    )
            
            # Registrar pago
            descripcion_planilla = f"Aplicación Planilla: {nombre_planilla}"
            if frecuencia:
                descripcion_planilla += f" — {frecuencia}"

            db_execute(
                conn,
                '''
                INSERT INTO pagos_prestamo (prestamo_id, monto, capital, interes, saldo_restante, descripcion, boleta_deposito, fecha, numero_comprobante)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (prestamo_id, monto, capital, interes, nuevo_saldo,
                 descripcion_planilla, boleta_deposito, fecha_pago, generar_numero_comprobante(conn))
            )

            if pago.get('detalle_id'):
                db_execute(conn, 'UPDATE planilla_masiva_detalles SET estado=?, monto=? WHERE id=?', ('aplicado', monto, pago['detalle_id']))
            
            procesados += 1
            total_capital += capital
            total_interes += interes
            resumen_aplicados.append({
                'numero': pago.get('numero', str(prestamo_id)),
                'monto': round(monto, 2),
                'capital': capital,
                'interes': interes,
            })
            
        except Exception as e:
            errores.append(f"Error procesando préstamo {pago.get('numero', prestamo_id)}: {str(e)}")
    
    if planilla_id and planilla:
        pendientes = db_fetchone(
            conn,
            "SELECT COUNT(*) FROM planilla_masiva_detalles WHERE planilla_id=? AND estado='pendiente'",
            (planilla_id,)
        )[0]
        estado_final = 'aplicada' if pendientes == 0 and procesados > 0 else ('parcial' if procesados > 0 else 'pendiente')
        db_execute(conn, '''
            UPDATE planillas_masivas
            SET estado=?, boleta_deposito=?, fecha_aplicacion=?, usuario_aplicacion=?
            WHERE id=?
        ''', (estado_final, boleta_deposito, datetime.now().isoformat(), session.get('username'), planilla_id))

    conn.commit()
    conn.close()

    log_auditoria_evento(
        modulo='prestamos',
        entidad='planilla_masiva',
        entidad_id=planilla_id,
        accion='aplicar',
        descripcion='Aplicación de pagos masivos de préstamos',
        datos={
            'procesados': procesados,
            'errores': len(errores),
            'boleta': boleta_deposito,
            'capital_total': round(total_capital, 2),
            'interes_total': round(total_interes, 2),
        }
    )
    
    return jsonify({
        'procesados': procesados,
        'errores': errores,
        'total': len(pagos),
        'planilla_id': planilla_id,
        'capital_total': round(total_capital, 2),
        'interes_total': round(total_interes, 2),
        'resumen_aplicados': resumen_aplicados,
    })
